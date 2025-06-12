import os
import qrcode
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image

def clean_filename(text):
    text = text.replace('*', 'x')
    return re.sub(r'[\\/:?"<>|]', '_', text)

# 檔案與工作表名稱
excel_path = 'input.xlsx'
sheet_name = '商品編號'

wb = load_workbook(excel_path)
ws = wb[sheet_name]

# 建立 output 資料夾
output_dir = 'output'
os.makedirs(output_dir, exist_ok=True)

# 從第3列開始讀取
for row_index, row in enumerate(ws.iter_rows(min_row=3), start=3):
    b_cell = row[1].value  # B欄 = index 1
    c_cell = row[2].value  # C欄 = index 2

    if b_cell is None or c_cell is None:
        continue

    b_str = str(b_cell).strip()
    c_str = str(c_cell).strip()
    combined = f"{b_str}@{c_str}"

    # 產生 QR Code
    qr_img = qrcode.make(combined)

    # 檔名清理
    safe_b = clean_filename(b_str)
    safe_c = clean_filename(c_str)
    filename = f"{safe_b} {safe_c}.png"
    save_path = os.path.join(output_dir, filename)

    # 儲存 QR code 圖片
    qr_img.save(save_path)

    # 插入到 Excel 的 A欄（index 0）
    excel_img = ExcelImage(save_path)
    excel_img.width = 100  # 可調整大小
    excel_img.height = 100
    cell_location = f"A{row_index}"
    ws.add_image(excel_img, cell_location)

    print(f"✅ 圖片儲存與插入: {save_path} -> {cell_location}")

# 儲存修改後的 Excel
wb.save(excel_path)
print("📄 Excel 已更新並儲存！")
